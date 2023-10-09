#!/usr/bin/python
# -*- coding: UTF-8 -*-

"""
Copyright (c) 2023, Cristóvão Beirão da Cruz e Silva <cristovao.silva@ist.utl.pt>

This software is provided 'as-is', without any express or implied warranty. In no event will the authors be held liable for any damages arising from the use of this software.

Permission is granted to anyone to use this software for any purpose, including commercial applications, and to alter it and redistribute it freely, subject to the following restrictions:

1. The origin of this software must not be misrepresented; you must not claim that you wrote the original software. If you use this software in a product, an acknowledgment in the product documentation would be appreciated but is not required.

2. Altered source versions must be plainly marked as such, and must not be misrepresented as being the original software.

3. This notice may not be removed or altered from any source distribution.
"""

from pathlib import Path
from xml.etree.ElementTree import ElementTree

class GeneOntology:

    _accession: str = ""
    _name: str = ""
    _namespace: str = ""
    _info : dict[str, str]

    def __init__(self, xml = None, accession: str = '', name: str = '', namespace: str = '', info = {}):
        self._info = {}

        if xml == None:
            if accession == "" or accession is None:
                raise RuntimeError("Trying to input a GO term with an empty accession")
            self._accession = accession
            self._name = name
            self._namespace = namespace
            self._info = info
        else:
            for entry in xml:  # TODO: Fix this to use the more recent xml format and to use ElementTree typing
                if entry.tag == 'id':
                    self._accession = entry.text
                elif entry.tag == 'name':
                    self._name = entry.text
                elif entry.tag == 'namespace':
                    self._namespace = entry.text
                elif entry.tag == 'def':
                    self._info['Definition'] = entry[0].text
                elif entry.tag == 'is_a':
                    if 'is a' not in self._info:
                        self._info['is a'] = []
                    self._info['is a'].append(entry.text)
                elif entry.tag == 'alt_id':
                    if 'Alt ID' not in self._info:
                        self._info['Alt ID'] = []
                    self._info['Alt ID'].append(entry.text)
                elif entry.tag == 'relationship':
                    if 'relationship' not in self._info:
                        self._info['relationship'] = {}
                    if entry[0].text not in self._info['relationship']:
                        self._info['relationship'][entry[0].text] = []
                    self._info['relationship'][entry[0].text].append(entry[1].text)
                elif entry.tag == 'comment' or entry.tag == 'replaced_by' or entry.tag == 'is_root':
                    self._info[entry.tag] = entry.text
                elif entry.tag == 'lexical_category' or entry.tag == 'disjoint_from' or entry.tag == 'union_of' or entry.tag == 'intersection_of' or entry.tag == 'subset' or entry.tag == 'synonym' or entry.tag == 'xref_analog' or entry.tag == 'consider' or entry.tag == 'is_obsolete' or entry.tag == 'is_anonymous':
                    continue
                else:
                    print(entry.tag, "is an unknown parameter.")

    @property
    def accession(self):
        """The accession property."""
        return self._accession

    @property
    def name(self):
        """The name property."""
        return self._name

    @property
    def namespace(self):
        """The namespace property."""
        return self._namespace

    @property
    def info(self):
        """The info property."""
        return self._info.copy()

    def hasGOTree(self, GO: str):
        GOLibrary = GOManager()
        if self._accession == GO:
            return True
        if "is a" in self._info:
            for Accession in self._info["is a"]:
                if GOLibrary[Accession].hasGOTree(GO):
                    return True
        if "Relationship" in self._info:
            if "part_of" in self._info["Relationship"] or "occurs_in" in self._info["Relationship"]:  # TODO: is this incomplete? it seems to be missing the occurs_in part
                if "part_of" in self._info["Relationship"]:
                    for Accession in self._info["Relationship"]["part_of"]:
                        if GOLibrary[Accession].hasGOTree(GO):
                            return True
        return False

    def hasGOTreeRelaxed(self, GO: str):
        GOLibrary = GOManager()
        if self._accession == GO:
            return True
        if "is a" in self._info:
            for Accession in self._info["is a"]:
                if GOLibrary[Accession].hasGOTreeRelaxed(GO):
                    return True
        if "Relationship" in self._info:
            if "part_of" in self._info["Relationship"] or "occurs_in" in self._info["Relationship"]:
                for Accession in self._info["Relationship"]["part_of"]:
                    if GOLibrary[Accession].hasGOTreeRelaxed(GO):
                        return True
        for Accession in GOLibrary.part_of(self._accession):
            if GOLibrary[Accession].hasGOTreeRelaxed(GO):
                return True
        return False

class GOManager:
    from io import TextIOWrapper

    _GOs: dict[str, GeneOntology] = {}
    _verbose: bool = False
    _cache_part_of: dict[str, list[str]] = {}

    def __init__(self, goOboFile: Path = None, verbose: bool = False):
        self._verbose = verbose

        if goOboFile is None:
            return

        if not goOboFile.exists() or not goOboFile.is_file():
            raise RuntimeError(f"You must define an existing file for the GO database: {goOboFile}")

        print("Loading Gene Ontologies")

        with goOboFile.open(mode='r') as file:
            line = file.readline()
            while line:
                if line[:-1] == '[Term]':
                    entry, line = self._loadGO(file)
                    self._GOs[entry.accession] = entry
                else:
                    line = file.readline()

        print("Finished loading Gene Ontologies")

    def keys(self):
        return self._GOs.keys()

    def items(self):
        return self._GOs.items()

    def values(self):
        return self._GOs.values()

    def __iter__(self):
        return iter(self._GOs)

    def __getitem__(self, accession: str) -> GeneOntology:
        try:
            return self._GOs[accession]
        except KeyError:
            for _, GO in self._GOs.items():
                GOInfo = GO.info
                if 'Alt ID' in GOInfo:
                    if accession in GOInfo['Alt ID']:
                        return GO
            raise KeyError("Unknown GO Accession number")

    def part_of(self, accession: str) -> list[str]:
        if len(self._cache_part_of) == 0:
            for goAcc, GO in self._GOs.items():
                GOInfo = GO.info
                if "Relationship" in GOInfo:
                    if "has_part" in GOInfo["Relationship"]:
                        for goAcc2 in GOInfo["Relationship"]["has_part"]:
                            if goAcc2 not in self._cache_part_of:
                                self._cache_part_of[goAcc2] = []
                            if goAcc not in self._cache_part_of[goAcc2]:
                                self._cache_part_of[goAcc2].append(goAcc)

        if accession in self._cache_part_of:
            return self._cache_part_of[accession]
        else:
            return []

    def _loadGO(self, file: TextIOWrapper) -> tuple[GeneOntology, str]:
        current_line = file.readline()
        Accession = ''
        Name = ''
        Namespace = ''
        Info = {}

        while current_line != "\n":
            if current_line[:3] == "id:":
                Accession = current_line[4:-1]
            elif current_line[:5] == "name:":
                Name = current_line[6:-1]
            elif current_line[:10] == "namespace:":
                Namespace = current_line[11:-1]
            elif current_line[:7] == "alt_id:":
                if "Alt ID" not in Info:
                    Info["Alt ID"] = []
                Info["Alt ID"].append(current_line[8:-1])
            elif current_line[:4] == "def:":
                Info["Definition"] = current_line[5:-1].split("\"")[1]
            elif current_line[:8] == "comment:":
                Info["Comment"] = current_line[9:-1]
            elif current_line[:5] == "is_a:":
                if "is a" not in Info:
                    Info["is a"] = []
                Info["is a"].append(current_line[6:-1].split(" ! ")[0])
            elif current_line[:16] == "intersection_of:":
                if "Intersection of" not in Info:
                    Info["Intersection of"] = []
                tmp = current_line[17:-1].split(" ! ")[0].split(" ")
                if len(tmp) == 2:
                    tmp = [tmp[1], tmp[0]]
                else:
                    tmp = [tmp[0]]
                Info["Intersection of"].append(tmp)
            elif current_line[:13] == "relationship:":
                if "Relationship" not in Info:
                    Info["Relationship"] = {}
                tmp = current_line[14:-1].split(" ! ")[0].split(" ")
                if tmp[0] not in Info["Relationship"]:
                    Info["Relationship"][tmp[0]] = []
                Info["Relationship"][tmp[0]].append(tmp[1])
            elif current_line[:12] == "is_obsolete:":
                if current_line[13:-1] == "true":
                    Info["Obsolete"] = True
            elif current_line[:12] == "replaced_by:":
                if "Replaced by" not in Info:
                    Info["Replaced by"] = []
                Info["Replaced by"].append(current_line[13:-1])
            elif current_line[:9] == "consider:":
                if "Consider" not in Info:
                    Info["Consider"] = []
                Info["Consider"].append(current_line[10:-1])
            elif current_line[:7] == "subset:":
                if "Subset" not in Info:
                    Info["Subset"] = []
                Info["Subset"].append(current_line[8:-1])
            elif current_line[:8] == "synonym:":
                pass  # TODO: implement synonyms
            elif current_line[:5] == "xref:":
                pass  # TODO: implement xrefs
            elif current_line[:11] == "created_by:":
                pass  # TODO: implement created_by
            elif current_line[:14] == "creation_date:":
                pass  # TODO: implement creation_date
            elif current_line[:15] == "property_value:":
                pass  # TODO: implement property_value
            elif current_line[:14] == "disjoint_from:":
                pass  # TODO: implement disjoint_from
            else:
                if self._verbose:
                    print(f"Unknown annotation line: {current_line}")
            current_line = file.readline()

        return GeneOntology(accession=Accession, name=Name, namespace=Namespace, info=Info), current_line

    def loadGOSlim(self, goSlimFile: Path, limitTo: str = ""):
        if not goSlimFile.exists() or not goSlimFile.is_file():
            raise RuntimeError(f"You must define an existing file for the GO database: {goSlimFile}")

        print(f"Loading Gene Ontology Slim: {goSlimFile}")

        with goSlimFile.open(mode='r') as file:
            line = file.readline()
            while(line):
                if line[:-1] == '[Term]':
                    entry, line = self._loadGO(file)

                    accession = entry.accession
                    if accession not in self._GOs:
                        for goAcc, GO in self._GOs.items():
                            if 'Alt ID' in GO.info:
                                if accession in GO.info['Alt ID']:
                                    accession = goAcc
                                    break

                    if accession not in self._GOs:
                        raise Exception("Unable to find the GO accession: "+accession+". Check if a newer base gene ontology is available.")

                    if "Subset" in entry.info:
                        if "Subset" not in self._GOs[accession].info:
                            self._GOs[accession]._info["Subset"] = []

                        for subset in entry.info["Subset"]:
                            if limitTo != "":
                                if subset != limitTo:
                                    continue
                            if subset not in self._GOs[accession].info["Subset"]:
                                self._GOs[accession]._info["Subset"].append(subset)
                else:
                    line = file.readline()

        print(f"Finished loading Gene Ontology Slim: {goSlimFile}")

class ProteinName:
    from xml.etree.ElementTree import Element

    def __init__(self, xml: Element):
        self.xml = xml  # TODO: parse the protein name info, see the uniprot xsd and https://www.uniprot.org/help/protein_names for a description

class GeneName:
    from xml.etree.ElementTree import Element

    _first_name: str = ""
    _primary_name: str = ""
    _synonyms: list[str]
    _ordered_locus: list[str]
    _orf: list[str]

    def __init__(self, xml: Element):
        self._synonyms = []
        self._ordered_locus = []
        self._orf = []

        for entry in xml:  # For format description, see the uniprot xsd and https://www.uniprot.org/help/gene_name
            if self._first_name == "":
                self._first_name = entry.text

            if 'evidence' in entry.attrib:  # TODO: what to do with the evidence attribute?
                pass

            if entry.attrib['type'] == 'primary':  # TODO: Add a check for multiple primaries?
                self._primary_name = entry.text
            elif entry.attrib['type'] == 'synonym':
                self._synonyms.append(entry.text)
            elif entry.attrib['type'] == 'ordered locus':
                self._ordered_locus.append(entry.text)
            elif entry.attrib['type'] == 'ORF':
                self._orf.append(entry.text)

    @property
    def name(self):
        if self._primary_name == "":
            return self._first_name
        return self._primary_name

class DBList:
    from xml.etree.ElementTree import Element

    _db_name: str = ""
    _references: dict[str, list[Element]]

    def __init__(self, dbName: str):
        self._db_name = dbName
        self._references = {}

    def keys(self):
        return self._references.keys()

    def items(self):
        return self._references.items()

    def values(self):
        return self._references.values()

    def __iter__(self):
        return iter(self._references)

    def __getitem__(self, accession: str) -> list[Element]:
        try:
            return self._references[accession]
        except KeyError:
            raise KeyError(f"Unknown accession for Database {self._db_name}: {accession}")

    def get(self, accession: str):
        return self[accession]

    def accessions(self):
        return list(self._references.keys())

    def add_reference(self, xml: Element):
        dbName = xml.attrib["type"]

        if dbName != self._db_name:
            raise RuntimeError(f"There was a serious problem with a mismatch of db names for the dbReferences: {dbName} vs {self._db_name}")

        accession = xml.attrib["id"]
        properties = []
        for child in xml:
            properties.append(child)

        if accession not in self._references:
            self._references[accession] = properties
        else:
            print(f"Maybe there is a clash of dbReferences - {dbName}:{accession}. Attempting to merge")
            self._references[accession] += properties

class DBReference:
    from xml.etree.ElementTree import Element

    _references: dict[str, DBList]

    def __init__(self):
        self._references = {}

    def keys(self):
        return self._references.keys()

    def items(self):
        return self._references.items()

    def values(self):
        return self._references.values()

    def __iter__(self):
        return iter(self._references)

    def __getitem__(self, DBname: str) -> DBList:
        try:
            return self._references[DBname]
        except KeyError:
            raise KeyError(f"Unknown Database: {DBname}")  # TODO: perhaps return empty list instead of throwing

    def get(self, DBname: str):
        return self[DBname]

    def add_reference(self, xml: Element):
        dbName = xml.attrib["type"]

        if dbName not in self._references:
            self._references[dbName] = DBList(dbName)

        self._references[dbName].add_reference(xml)

class Protein:
    from xml.etree.ElementTree import Element

    _xml: Element = None
    _accession: str = ""
    _secondary_accessions: list[str]
    _name: str = ""  # Name of the entry in UP, not necessarily a well known name: https://www.uniprot.org/help/entry_name
    _protein_name: ProteinName = None
    _gene_name: GeneName = None
    # others missing here
    _db_reference: DBReference
    _protein_existence: str = ""

    def __init__(self, xml: Element):
        self._xml = xml
        self._secondary_accessions = []
        self._db_reference = DBReference()

        # TODO: Enforce minimums and maximums from XSD file
        for entry in self._xml:
            if entry.tag == "{http://uniprot.org/uniprot}accession":
                # TODO: validate accessions with RE: [OPQ][0-9][A-Z0-9]{3}[0-9]|[A-NR-Z][0-9]([A-Z][A-Z0-9]{2}[0-9]){1,2} (from https://www.uniprot.org/help/accession_numbers)
                if self._accession == "":
                    self._accession = entry.text
                else:
                    self._secondary_accessions.append(entry.text)
            elif entry.tag == "{http://uniprot.org/uniprot}name":
                if self._name == "":
                    self._name = entry.text
                else:
                    print(f"A protein with multiple names was found: {self._accession}")
            elif entry.tag == "{http://uniprot.org/uniprot}protein":
                self._protein_name = ProteinName(entry)
            elif entry.tag == "{http://uniprot.org/uniprot}gene":
                self._gene_name = GeneName(entry)
            elif entry.tag == "{http://uniprot.org/uniprot}organism":
                pass  # TODO: implement a separate DB for the organisms (in the ProteinManager) and store only the Taxonomy Accession here
            elif entry.tag == "{http://uniprot.org/uniprot}organismHost":
                pass  # TODO: implement this
            elif entry.tag == "{http://uniprot.org/uniprot}geneLocation":
                pass  # TODO: implement this
            elif entry.tag == "{http://uniprot.org/uniprot}reference":
                pass  # TODO: implement this
            elif entry.tag == "{http://uniprot.org/uniprot}comment":
                pass  # TODO: implement this
            elif entry.tag == "{http://uniprot.org/uniprot}dbReference":
                self._db_reference.add_reference(entry)
            elif entry.tag == "{http://uniprot.org/uniprot}proteinExistence":
                self._protein_existence = entry.attrib["type"]  # TODO: add check for extra (should not happen)
            elif entry.tag == "{http://uniprot.org/uniprot}keyword":
                pass  # TODO: implement this (consider using a keyword DB, fetching from UP the full list of keywords)
            elif entry.tag == "{http://uniprot.org/uniprot}feature":
                pass  # TODO: implement this
            elif entry.tag == "{http://uniprot.org/uniprot}evidence":
                pass  # TODO: implement this
            elif entry.tag == "{http://uniprot.org/uniprot}sequence":
                pass  # TODO: implement this
            else:
                print(entry.tag)
                break

    @property
    def name(self):
        return self._name

    @property
    def gene_name(self):
        return self._gene_name.name

    @property
    def accession(self):
        return self._accession

    @property
    def secondary_accessions(self):
        return self._secondary_accessions.copy()

    @property
    def db_references(self):
        return self._db_reference

    def has_accession(self, accession):
        if accession == self._accession:
            return True
        elif accession in self._secondary_accessions:
            return True
        return False

class ProteinManager:

    _proteins: dict[str, Protein] = {}

    def __init__(self, proteinXMLFile: Path = None, verbose: bool = False):

        if not proteinXMLFile.exists() or not proteinXMLFile.is_file():
            raise RuntimeError(f"You must define an existing file for the result of the protein XML query: {proteinXMLFile}")

        tree = ElementTree()
        tree.parse(proteinXMLFile)
        root = tree.getroot()

        for element in root:
            if element.tag == '{http://uniprot.org/uniprot}entry':
                protein = Protein(element)
                self._proteins[protein.accession] = protein

    def keys(self):
        return self._proteins.keys()

    def items(self):
        return self._proteins.items()

    def values(self):
        return self._proteins.values()

    def __iter__(self):
        return iter(self._proteins)

    def __getitem__(self, accession: str) -> Protein:
        try:
            return self._proteins[accession]
        except KeyError:
            for protAcc, protein in self._proteins.items():
                pass # TODO: search in the alternative accessions
            raise KeyError(f"Unknown protein Accession number: {accession}")

def script_main(
        goOboPath: Path,
        dataPaths: list[Path],
        goSlim: str = "goslim_generic",
        goNamespace: str = 'A',  # Options: B, M, C, A, None
    ):
    if goNamespace not in ['B', 'M', 'C', 'A', None]:
        raise RuntimeError(f'Invalid GO Namespace filter selected: {goNamespace}')

    namespaces_to_run = ["B", "M", "C"]  # Option A - All
    if goNamespace is None:
        namespaces_to_run = []
    elif goNamespace != 'A':
        namespaces_to_run = [goNamespace]

    GOM = GOManager(goOboFile = goOboPath/"go.obo")

    for basePath in dataPaths:
        PM = ProteinManager(proteinXMLFile = basePath/"listUP.xml")

        from openpyxl import Workbook
        wb = Workbook()

        info_sheet = wb.active
        info_sheet.title = "Info"

        info_sheet.cell(row=1, column=1, value="ID")
        info_sheet.cell(row=1, column=2, value="Accession")
        info_sheet.cell(row=1, column=3, value="Name")
        info_sheet.cell(row=1, column=4, value="Gene Name")
        info_sheet.cell(row=1, column=5, value="Molecular Function")
        info_sheet.cell(row=1, column=6, value="Biological Process")
        info_sheet.cell(row=1, column=7, value="Cellular Component")

        current_row = 2
        for protAcc in PM:
            info_sheet.cell(row = current_row, column=1, value=protAcc)
            info_sheet.cell(row = current_row, column=2, value=protAcc)
            info_sheet.cell(row = current_row, column=3, value=PM[protAcc].name)
            info_sheet.cell(row = current_row, column=4, value=PM[protAcc].gene_name)

            m_counter = 0
            b_counter = 0
            c_counter = 0

            if "GO" in PM[protAcc].db_references:
                for goAcc in PM[protAcc].db_references["GO"]:
                    goEntry = GOM[goAcc]
                    if goEntry.namespace == "molecular_function":
                        info_sheet.cell(row = current_row + m_counter, column = 5, value=goEntry.name)
                        m_counter += 1
                    elif goEntry.namespace == "biological_process":
                        info_sheet.cell(row = current_row + b_counter, column = 6, value=goEntry.name)
                        b_counter += 1
                    elif goEntry.namespace == "cellular_component":
                        info_sheet.cell(row = current_row + c_counter, column = 7, value=goEntry.name)
                        c_counter += 1
                    else:
                        print(f"Unknown namespace: {goEntry.namespace}")

            current_row += max(1, m_counter, b_counter, c_counter)

        wb.save(basePath/"SummaryGO.xlsx")

        for goNS in namespaces_to_run:
            if goNS == "M":
                goNS = "molecular_function"
            elif goNS == "B":
                goNS = "biological_process"
            elif goNS == "C":
                goNS = "cellular_component"

            considerGoAcc: dict[str, list[str]] = {}
            for GOAcc in GOM:
                if goNS != GOM[GOAcc].namespace:
                    continue
                if "Subset" in GOM[GOAcc].info:
                    if goSlim in GOM[GOAcc].info["Subset"]:
                        considerGoAcc[GOAcc] = []

            for protAcc in PM:
                for GOAcc in considerGoAcc:
                    foundGO = False

                    if "GO" in PM[protAcc].db_references:
                        for entry in PM[protAcc].db_references["GO"]:
                            if GOM[entry].hasGOTree(GOAcc):
                                foundGO = True
                                break

                    if foundGO:
                        considerGoAcc[GOAcc].append(protAcc)

            removeAcc = ["GO:0008150", "GO:0003674", "GO:0005575"]  # These are the root accessions
            for GOAcc, protList in considerGoAcc.items():
                count = len(protList)
                if count == 0:
                    removeAcc.append(GOAcc)
            for GOAcc in removeAcc:
                if GOAcc in considerGoAcc:
                    del considerGoAcc[GOAcc]

            wb = Workbook()

            info_sheet = wb.active
            info_sheet.title = "Info"

            info_sheet.cell(row=1, column=1, value="GO Accession")
            info_sheet.cell(row=1, column=2, value="GO Name")
            info_sheet.cell(row=1, column=3, value="Protein Count")
            info_sheet.cell(row=1, column=4, value="Protein Accession")
            info_sheet.cell(row=1, column=5, value="Protein Entry Name")
            info_sheet.cell(row=1, column=6, value="Gene Name")

            current_row = 2
            for GOAcc, protList in considerGoAcc.items():
                info_sheet.cell(row=current_row, column=1, value=GOAcc)
                info_sheet.cell(row=current_row, column=2, value=GOM[GOAcc].name)
                info_sheet.cell(row=current_row, column=3, value=len(protList))

                offset = 0
                for protAcc in protList:
                    info_sheet.cell(row=current_row + offset, column=4, value=protAcc)
                    info_sheet.cell(row=current_row + offset, column=5, value=PM[protAcc].name)
                    info_sheet.cell(row=current_row + offset, column=6, value=PM[protAcc].gene_name)
                    offset += 1

                current_row += max(1, len(protList))

            wb.save(basePath/f"Summary_{goNS}.xlsx")

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
                    prog='go_ana.py',
                    description='This script performs the GO analysis',
                    #epilog='Text at the bottom of help'
                    )

    parser.add_argument(
        '-g',
        '--goOboPath',
        metavar = 'PATH',
        type = Path,
        help = 'Path to the directory cotaining the go.obo file with the GO database to use as a reference',
        required = True,
        dest = 'goOboPath',
    )
    parser.add_argument(
        '-d',
        '--dataPath',
        metavar = 'PATH',
        type = Path,
        help = 'Path to the directory cotaining listUP.xml input data file and where to store the output',  # TODO: add possibility to give multiple
        required = True,
        dest = 'dataPath',
    )
    parser.add_argument(
        '-n',
        '--namespace',
        metavar = 'NAMESPACE',
        type = str,
        help = "Namespace of GO to consider (default A): B -> Biological Process; M -> Molecula Function; C -> Cellular Component; A -> All",
        choices = ['B', 'M', 'C', 'A'],
        default = 'A',
        dest = 'goNamespace',
    )

    args = parser.parse_args()

    goOboPath: Path = args.goOboPath
    if not goOboPath.exists() or not goOboPath.is_dir():
        raise RuntimeError("You must define an existing Path for goOboPath")
    if not (goOboPath/'go.obo').exists() or not (goOboPath/'go.obo').is_file():
        raise RuntimeError("You must specify a path for goOboPath which contains the go.obo file (download from https://geneontology.org/)")
    goOboPath = goOboPath.absolute()

    dataPath: Path = args.dataPath
    if not dataPath.exists() or not dataPath.is_dir():
        raise RuntimeError("You must define an existing Path for dataPath")
    if not (dataPath/'listUP.xml').exists() or not (dataPath/'listUP.xml').is_file():
        raise RuntimeError("You must specify a path for dataPath which contains the listUP.xml file (the results of a query to https://www.uniprot.org/)")
    dataPath = dataPath.absolute()

    script_main(
        goOboPath = goOboPath,
        dataPaths = [dataPath],
        goNamespace = args.goNamespace,
        )
